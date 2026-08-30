
"""Diagnostic : à lancer en local sur ton vrai fichier XTB pour voir la
structure exacte (noms d'onglets réels + 5 premières lignes brutes)."""
 
import sys
 
import openpyxl
 
 
def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("usage: python debug_xtb.py <fichier.xlsx> [nom_onglet]")
 
    path = sys.argv[1]
 
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print("Onglets trouvés :")
    for name in sheet_names:
        print(f"  - {name!r}")
    wb.close()
 
    target = sys.argv[2] if len(sys.argv) > 2 else "CLOSED POSITION HISTORY"
    print(f"\nInspection de l'onglet : {target!r} (mode read_only=True)")
 
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb[target]
    print(f"\n5 premières lignes brutes de {target!r} :")
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= 5:
            break
        types = [type(c).__name__ for c in row]
        print(f"  ligne {i}: {row}")
        print(f"    types: {types}")
    wb.close()
 
    print(f"\n--- Même onglet, mode read_only=False (fallback si vide ci-dessus) ---")
    wb2 = openpyxl.load_workbook(path, read_only=False, data_only=True)
    sheet2 = wb2[target]
    print(f"dimensions déclarées: {sheet2.dimensions}, max_row={sheet2.max_row}, max_column={sheet2.max_column}")
    for i, row in enumerate(sheet2.iter_rows(values_only=True)):
        non_empty = any(v is not None for v in row)
        marker = "  <-- contient des valeurs" if non_empty else ""
        print(f"  ligne {i}: {row}{marker}")
    wb2.close()
 
 
if __name__ == "__main__":
    main()