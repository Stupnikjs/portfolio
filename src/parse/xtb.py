"""Équivalent Python de pf_extract::xtb (positions ouvertes/fermées XTB)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl


from ..schema import Asset, AssetIdentifiers, AssetKind, Platform, Transaction, TransactionKind


@dataclass
class XtbClosedPosition:
    position_id: str
    symbol: str
    side: str
    volume: float
    open_time: datetime
    open_price: float
    close_time: datetime
    close_price: float
    open_origin: str
    close_origin: str
    purchase_value: float
    sale_value: float
    sl: Optional[float]
    tp: Optional[float]
    margin: Optional[float]
    commission: float
    swap: float
    rollover: float
    gross_pl: float
    source_file: str

    def to_transactions(self) -> list[Transaction]:
        currency = _infer_currency(self.symbol)
        asset = Asset(symbol=self.symbol, name=self.symbol,
                  kind=AssetKind.STOCK, ref_currency=currency,
                  identifiers=AssetIdentifiers())
        return [
            Transaction(
                platform=Platform.XTB, account_label="XTB", kind=TransactionKind.BUY,
                asset=asset, quantity=self.volume, price=self.open_price,
                amount=self.purchase_value, quote_currency=None, time=self.open_time,
                value_eur=self.purchase_value,
                external_id=f"{self.position_id}-buy", remark=None, source_file=self.source_file,
            ),
            Transaction(
                platform=Platform.XTB, account_label="XTB", kind=TransactionKind.SELL,
                asset=asset, quantity=self.volume, price=self.close_price,
                amount=self.sale_value, quote_currency=None, time=self.close_time,
                value_eur=self.sale_value,
                external_id=f"{self.position_id}-sell", remark=None, source_file=self.source_file,
            ),
        ]


@dataclass
class XtbOpenPosition:
    position_id: str
    symbol: str
    side: str
    volume: float
    open_time: datetime
    open_price: float
    market_price: float
    purchase_value: float
    sl: Optional[float]
    tp: Optional[float]
    margin: Optional[float]
    commission: float
    swap: float
    rollover: float
    gross_pl: float
    comment: Optional[str]
    source_file: str

    def to_transaction(self) -> Transaction:
        currency = _infer_currency(self.symbol)
        asset = Asset(symbol=self.symbol, name=self.symbol,
                  kind=AssetKind.STOCK, ref_currency=currency,
                  identifiers=AssetIdentifiers())
        return Transaction(
            platform=Platform.XTB, account_label="XTB", kind=TransactionKind.BUY,
            asset=asset, quantity=self.volume, price=self.open_price,
            amount=self.purchase_value, quote_currency=None, time=self.open_time,
            value_eur=self.purchase_value,
            external_id=self.position_id, remark=self.comment, source_file=self.source_file,
        )


def _infer_currency(symbol: str) -> str:
    suffix = symbol.rsplit(".", 1)[-1] if "." in symbol else ""
    return {
        "DE": "EUR", "NL": "EUR", "FR": "EUR", "ES": "EUR", "IT": "EUR",
        "UK": "GBP", "US": "USD",
    }.get(suffix, "EUR")


def find_sheet_by_prefix(path: Path, prefix: str) -> str:
    """Trouve dynamiquement le nom d'onglet commençant par un préfixe donné,
    utile pour 'OPEN POSITION <date>' dont le nom exact change à chaque export.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        for name in wb.sheetnames:
            if name.startswith(prefix):
                return name
    finally:
        wb.close()
    raise ValueError(f"aucun onglet commençant par '{prefix}'")


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_f64(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _cell_datetime(value) -> datetime:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, str):
        naive = datetime.strptime(value.strip(), "%d/%m/%Y %H:%M:%S")
        return naive.replace(tzinfo=timezone.utc)
    raise ValueError(f"format de date inattendu: {value!r}")


def _find_header_row(rows: list[tuple], required_col: str) -> int:
    """Trouve l'index de la ligne d'en-tête -- celle qui contient
    `required_col`. Les exports XTB ont un nombre variable de lignes de
    titre/résumé de compte au-dessus du vrai tableau, et ce nombre change
    d'une version d'export à l'autre -- chercher par nom plutôt que
    supposer une position fixe évite de recasser au prochain changement
    de format XTB."""
    for i, row in enumerate(rows):
        if row and required_col in row:
            return i
    raise ValueError(f"en-tête introuvable (colonne '{required_col}' non trouvée)")


def _column_map(header_row: tuple) -> dict[str, int]:
    return {str(name).strip(): i for i, name in enumerate(header_row) if name is not None}


def parse_closed_positions(path: Path, sheet_name: str) -> list[XtbClosedPosition]:
    """Parse l'onglet 'Closed Positions' d'un export XTB xlsx.

    NB: pas de `read_only=True` ici -- openpyxl s'appuie en mode read_only sur
    la balise `<dimension>` du XML pour savoir quelle plage lire, et les
    exports XTB ont cette balise absente/incorrecte (symptôme observé :
    iter_rows ne renvoie que des tuples vides). Le mode normal l'ignore et lit
    vraiment la feuille.
    """
    source_file = str(path)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = _find_header_row(rows, "Position ID")
        col = _column_map(rows[header_idx])
        out: list[XtbClosedPosition] = []

        for row in rows[header_idx + 1:]:
            position_id = _cell_f64(row[col["Position ID"]])
            if position_id is None:
                continue  # ligne "Total" ou vide

            out.append(XtbClosedPosition(
                position_id=str(int(position_id)),
                symbol=_cell_str(row[col["Ticker"]]),
                side=_cell_str(row[col["Type"]]),
                volume=_cell_f64(row[col["Volume"]]) or 0.0,
                open_time=_cell_datetime(row[col["Open Time (UTC)"]]),
                open_price=_cell_f64(row[col["Open Price"]]) or 0.0,
                close_time=_cell_datetime(row[col["Close Time (UTC)"]]),
                close_price=_cell_f64(row[col["Close Price"]]) or 0.0,
                open_origin=_cell_str(row[col["Open Origin"]]) if "Open Origin" in col else "",
                close_origin=_cell_str(row[col["Close Origin"]]) if "Close Origin" in col else "",
                purchase_value=_cell_f64(row[col["Purchase Value"]]) or 0.0,
                sale_value=_cell_f64(row[col["Sale Value"]]) or 0.0,
                sl=_cell_f64(row[col["Stop Loss"]]),
                tp=_cell_f64(row[col["Take Profit"]]),
                margin=_cell_f64(row[col["Margin"]]),
                commission=_cell_f64(row[col["Commission"]]) or 0.0,
                swap=_cell_f64(row[col["Swap"]]) or 0.0,
                rollover=_cell_f64(row[col["Rollover"]]) or 0.0,
                gross_pl=_cell_f64(row[col["Gross Profit"]]) or 0.0,
                source_file=source_file,
            ))
        return out
    finally:
        wb.close()


def parse_open_positions(path: Path, sheet_name: str) -> list[XtbOpenPosition]:
    """Parse l'onglet 'Open Positions' d'un export XTB xlsx.

    Le tableau contient deux types de lignes : des lignes "résumé" par
    instrument (colonne Type vide, Instrument/Position = nom lisible) et des
    lignes "détail" par position individuelle (Type='BUY', Instrument/Position
    = l'identifiant numérique de la position). Seules les secondes sont
    retenues.
    """
    source_file = str(path)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = _find_header_row(rows, "Instrument/Position")
        col = _column_map(rows[header_idx])
        out: list[XtbOpenPosition] = []

        for row in rows[header_idx + 1:]:
            side = _cell_str(row[col["Type"]])
            if not side:
                continue
            position_id = _cell_str(row[col["Instrument/Position"]])
            if not position_id:
                continue

            out.append(XtbOpenPosition(
                position_id=position_id,
                symbol=_cell_str(row[col["Ticker"]]),
                side=side,
                volume=_cell_f64(row[col["Volume"]]) or 0.0,
                open_time=_cell_datetime(row[col["Open time (UTC)"]]),
                open_price=_cell_f64(row[col["Open price"]]) or 0.0,
                market_price=_cell_f64(row[col["Current price"]]) or 0.0,
                # Value = Volume x Current price (valeur au moment du rapport,
                # pas coût d'acquisition) -- choix assumé, pas Volume x Open price.
                purchase_value=_cell_f64(row[col["Value"]]) or 0.0,
                sl=_cell_f64(row[col["Stop Loss"]]),
                tp=_cell_f64(row[col["Take Profit"]]),
                margin=_cell_f64(row[col["Margin"]]),
                commission=_cell_f64(row[col["Open Commission"]]) or 0.0,
                swap=_cell_f64(row[col["Swap"]]) or 0.0,
                rollover=_cell_f64(row[col["Rollover"]]) or 0.0,
                gross_pl=_cell_f64(row[col["Gross Profit"]]) or 0.0,
                comment=None,  # colonne "Comment" absente du nouvel export
                source_file=source_file,
            ))
        return out
    finally:
        wb.close()


# Mapping des types d'opération 'Cash Operations' -> TransactionKind. Si un
# nouveau type apparaît dans un futur export, parse_cash_operations lève une
# ValueError explicite plutôt que de l'ignorer silencieusement -- l'ajouter
# ici quand il se présente.
_CASH_KIND_MAP: dict[str, TransactionKind] = {
    "Deposit": TransactionKind.DEPOSITE,
    "Free funds interest": TransactionKind.DEPOSITE,
    "Stock purchase": TransactionKind.WITHDRAW,     # jambe cash de l'achat -- le titre est déjà géré par parse_open_positions/parse_closed_positions
    "Stock sell": TransactionKind.DEPOSITE,         # jambe cash de la vente -- idem
    "Dividend": TransactionKind.DEPOSITE,
    "Withholding tax": TransactionKind.FEE,         # retenue à la source sur dividende
    "Tax IFTT": TransactionKind.FEE,                # taxe française sur transactions financières
    "Fractional shares": TransactionKind.DEPOSITE,  # compensation cash pour rompus d'actions
}

# Ignoré : mouvement interne entre sous-comptes XTB, pas une entrée/sortie
# réelle de patrimoine -- le cash reste détenu, juste ailleurs. Vu sur les
# deux comptes (négatif côté principal "transfer out", positif côté PEA
# "transfer in").
_SKIP_CASH_TYPES = {"PEA deposit"}


def parse_cash_operations(path: Path, sheet_name: str) -> list[Transaction]:
    """Parse l'onglet 'Cash Operations' d'un export XTB xlsx en transactions
    de cash EUR (dépôts, dividendes, taxes, achats/ventes de titres...). Les
    virements internes entre sous-comptes XTB (_SKIP_CASH_TYPES) sont ignorés."""
    source_file = str(path)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = _find_header_row(rows, "ID")
        col = _column_map(rows[header_idx])
        out: list[Transaction] = []

        eur_asset = Asset(symbol="EUR", name="EUR", kind=AssetKind.CASH, ref_currency="EUR", identifiers=AssetIdentifiers())

        for row in rows[header_idx + 1:]:
            op_type = _cell_str(row[col["Type"]])
            if not op_type or op_type == "Total":
                continue
            if op_type in _SKIP_CASH_TYPES:
                continue

            kind = _CASH_KIND_MAP.get(op_type)
            if kind is None:
                raise ValueError(f"Type d'opération Cash non géré: {op_type!r}")

            amount = _cell_f64(row[col["Amount"]]) or 0.0
            op_id = _cell_str(row[col["ID"]])

            out.append(Transaction(
                platform=Platform.XTB,
                account_label="XTB",
                kind=kind,
                asset=eur_asset,
                quantity=abs(amount),
                price=1.0,
                value_eur=abs(amount),
                amount=amount,
                quote_currency="EUR",
                time=_cell_datetime(row[col["Time"]]),
                external_id=f"xtb-cash-{op_id}",
                remark=_cell_str(row[col["Comment"]]) or None,
                source_file=source_file,
            ))
        return out
    finally:
        wb.close()