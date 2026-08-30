"""Équivalent Python de pf_extract::xtb (positions ouvertes/fermées XTB)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

import openpyxl

from .registry import AssetRegistry
from .schema import AssetIdentifiers, AssetKind, Platform, Transaction, TransactionKind


@dataclass
class XtbClosedPosition:
    position_id: str
    symbol: str
    side: str
    volume: float
    open_time: datetime
    open_price: float
    close_time: datetime
    close_price: float
    open_origin: str
    close_origin: str
    purchase_value: float
    sale_value: float
    sl: Optional[float]
    tp: Optional[float]
    margin: Optional[float]
    commission: float
    swap: float
    rollover: float
    gross_pl: float
    source_file: str

    def to_transactions(self, registry: AssetRegistry) -> list[Transaction]:
        currency = _infer_currency(self.symbol)
        asset_id = registry.find_or_create(
            self.symbol, self.symbol, AssetKind.STOCK, currency, AssetIdentifiers(),
        )
        return [
            Transaction(
                platform=Platform.XTB, account_label="XTB", kind=TransactionKind.BUY,
                asset_id=asset_id, quantity=self.volume, price=self.open_price,
                amount=self.purchase_value, quote_currency=None, time=self.open_time,
                external_id=self.position_id, remark=None, source_file=self.source_file,
            ),
            Transaction(
                platform=Platform.XTB, account_label="XTB", kind=TransactionKind.SELL,
                asset_id=asset_id, quantity=self.volume, price=self.close_price,
                amount=self.sale_value, quote_currency=None, time=self.close_time,
                external_id=self.position_id, remark=None, source_file=self.source_file,
            ),
        ]


@dataclass
class XtbOpenPosition:
    position_id: str
    symbol: str
    side: str
    volume: float
    open_time: datetime
    open_price: float
    market_price: float
    purchase_value: float
    sl: Optional[float]
    tp: Optional[float]
    margin: Optional[float]
    commission: float
    swap: float
    rollover: float
    gross_pl: float
    comment: Optional[str]
    source_file: str

    def to_transaction(self, registry: AssetRegistry) -> Transaction:
        currency = _infer_currency(self.symbol)
        asset_id = registry.find_or_create(
            self.symbol, self.symbol, AssetKind.STOCK, currency, AssetIdentifiers(),
        )
        return Transaction(
            platform=Platform.XTB, account_label="XTB", kind=TransactionKind.BUY,
            asset_id=asset_id, quantity=self.volume, price=self.open_price,
            amount=self.purchase_value, quote_currency=None, time=self.open_time,
            external_id=self.position_id, remark=self.comment, source_file=self.source_file,
        )


def _infer_currency(symbol: str) -> str:
    suffix = symbol.rsplit(".", 1)[-1] if "." in symbol else ""
    return {
        "DE": "EUR", "NL": "EUR", "FR": "EUR", "ES": "EUR", "IT": "EUR",
        "UK": "GBP", "US": "USD",
    }.get(suffix, "EUR")


def find_sheet_by_prefix(path: Path, prefix: str) -> str:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        for name in wb.sheetnames:
            if name.startswith(prefix):
                return name
    finally:
        wb.close()
    raise ValueError(f"aucun onglet commençant par '{prefix}'")


def _row_cell(row: tuple, idx: int):
    """Accède à un index de ligne en sécurité (évite IndexError si la ligne est trop courte)."""
    return row[idx] if len(row) > idx else None


def _cell_str(value) -> str:
    """Équivalent de cell_string en Rust. Gère les nombres flottants sans .0"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_f64(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _cell_datetime(value) -> datetime:
    """Équivalent de cell_datetime. Gère les objets datetime, strings, et numéros Excel."""
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, str):
        naive = datetime.strptime(value.strip(), "%d/%m/%Y %H:%M:%S")
        return naive.replace(tzinfo=timezone.utc)
    if isinstance(value, (int, float)):
        # Conversion du format de date Excel (nombre de jours depuis 1899-12-30)
        naive = datetime(1899, 12, 30) + timedelta(days=float(value))
        return naive.replace(tzinfo=timezone.utc)
    raise ValueError(f"format de date inattendu: {value!r}")


def parse_closed_positions(path: Path, sheet_name: str) -> list[XtbClosedPosition]:
    """Parse l'onglet 'Closed Position History' d'un export XTB xlsx."""
    source_file = str(path)
    # On enlève read_only=True pour s'assurer que toutes les colonnes sont lues
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        sheet = wb[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        next(rows, None)  # en-tête
        out: list[XtbClosedPosition] = []

        for row in rows:
            position_id_val = _row_cell(row, 0)
            position_id = _cell_f64(position_id_val)
            if position_id is None:
                continue
            if position_id == 0.0 and _row_cell(row, 1) is None:
                continue

            out.append(XtbClosedPosition(
                position_id=str(int(position_id)),
                symbol=_cell_str(_row_cell(row, 1)),
                side=_cell_str(_row_cell(row, 2)),
                volume=_cell_f64(_row_cell(row, 3)) or 0.0,
                open_time=_cell_datetime(_row_cell(row, 4)),
                open_price=_cell_f64(_row_cell(row, 5)) or 0.0,
                close_time=_cell_datetime(_row_cell(row, 6)),
                close_price=_cell_f64(_row_cell(row, 7)) or 0.0,
                open_origin=_cell_str(_row_cell(row, 8)),
                close_origin=_cell_str(_row_cell(row, 9)),
                purchase_value=_cell_f64(_row_cell(row, 10)) or 0.0,
                sale_value=_cell_f64(_row_cell(row, 11)) or 0.0,
                sl=_cell_f64(_row_cell(row, 12)),
                tp=_cell_f64(_row_cell(row, 13)),
                margin=_cell_f64(_row_cell(row, 14)),
                commission=_cell_f64(_row_cell(row, 15)) or 0.0,
                swap=_cell_f64(_row_cell(row, 16)) or 0.0,
                rollover=_cell_f64(_row_cell(row, 17)) or 0.0,
                gross_pl=_cell_f64(_row_cell(row, 18)) or 0.0,
                source_file=source_file,
            ))
        return out
    finally:
        wb.close()


def parse_open_positions(path: Path, sheet_name: str) -> list[XtbOpenPosition]:
    """Parse l'onglet 'OPEN POSITION <date>' d'un export XTB xlsx."""
    source_file = str(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        sheet = wb[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        next(rows, None)  # en-tête
        out: list[XtbOpenPosition] = []

        for row in rows:
            position_id_val = _row_cell(row, 0)
            position_id = _cell_f64(position_id_val)
            if position_id is None:
                continue
            if position_id == 0.0 and _row_cell(row, 1) is None:
                continue

            comment_str = _cell_str(_row_cell(row, 15))
            comment = comment_str if comment_str else None

            out.append(XtbOpenPosition(
                position_id=str(int(position_id)),
                symbol=_cell_str(_row_cell(row, 1)),
                side=_cell_str(_row_cell(row, 2)),
                volume=_cell_f64(_row_cell(row, 3)) or 0.0,
                open_time=_cell_datetime(_row_cell(row, 4)),
                open_price=_cell_f64(_row_cell(row, 5)) or 0.0,
                market_price=_cell_f64(_row_cell(row, 6)) or 0.0,
                purchase_value=_cell_f64(_row_cell(row, 7)) or 0.0,
                sl=_cell_f64(_row_cell(row, 8)),
                tp=_cell_f64(_row_cell(row, 9)),
                margin=_cell_f64(_row_cell(row, 10)),
                commission=_cell_f64(_row_cell(row, 11)) or 0.0,
                swap=_cell_f64(_row_cell(row, 12)) or 0.0,
                rollover=_cell_f64(_row_cell(row, 13)) or 0.0,
                gross_pl=_cell_f64(_row_cell(row, 14)) or 0.0,
                comment=comment,
                source_file=source_file,
            ))
        return out
    finally:
        wb.close()